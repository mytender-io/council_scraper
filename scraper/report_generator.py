"""
Report Generator Module

Creates Excel reports from processed premises licence data.
Generates weekly reports with insights, trends, and business intelligence.
"""

import json
from typing import List, Dict, Optional, Any, Tuple
from datetime import datetime, timedelta
from pathlib import Path
import logging

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

from .models import PremisesLicence, WeeklyReport, LicenceType, LicenceStatus
from .config import get_settings

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates Excel reports from premises licence data"""
    
    def __init__(self):
        self.settings = get_settings()
        
    def generate_weekly_report(self, licences: List[PremisesLicence], week_start: Optional[datetime] = None) -> str:
        """Generate comprehensive weekly report"""
        if week_start is None:
            # Default to current week (Monday start)
            today = datetime.now().date()
            week_start = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
        
        week_end = week_start + timedelta(days=7)
        
        logger.info(f"Generating weekly report for {week_start.date()} to {week_end.date()}")
        
        # Filter licences for the week
        weekly_licences = self._filter_licences_for_period(licences, week_start, week_end)
        
        # Generate report data structure
        report_data = WeeklyReport(
            report_date=datetime.now(),
            period_start=week_start,
            period_end=week_end,
            total_licences=len(licences),
            new_licences=len(weekly_licences),
            councils_scraped=len(set(licence.council_name for licence in licences)),
            councils_successful=len(set(licence.council_name for licence in licences if licence.council_name)),
            summary_stats=self._calculate_summary_stats(licences, weekly_licences),
            licences_by_type=self._count_by_type(weekly_licences),
            licences_by_status=self._count_by_status(weekly_licences),
            top_councils=self._get_top_councils(weekly_licences),
            error_summary=[]  # Would need error data from scraping results
        )
        
        # Create Excel workbook
        filename = self._create_excel_report(report_data, licences, weekly_licences)
        
        logger.info(f"Weekly report generated: {filename}")
        return filename
    
    def generate_full_dataset_report(self, licences: List[PremisesLicence]) -> str:
        """Generate comprehensive report of all licence data"""
        logger.info(f"Generating full dataset report for {len(licences)} licences")
        
        # Create Excel workbook with multiple sheets
        filename = f"{self.settings.reports_dir}/full_dataset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Main data sheet
            self._create_main_data_sheet(writer, licences)
            
            # Summary sheet
            self._create_summary_sheet(writer, licences)
            
            # Business intelligence sheets
            self._create_council_analysis_sheet(writer, licences)
            self._create_business_type_analysis_sheet(writer, licences)
            self._create_activity_analysis_sheet(writer, licences)
            self._create_geographic_analysis_sheet(writer, licences)
            
            # Raw data sheet (for power users)
            self._create_raw_data_sheet(writer, licences)
        
        # Post-process workbook for formatting
        self._format_workbook(filename)
        
        logger.info(f"Full dataset report generated: {filename}")
        return filename
    
    def _filter_licences_for_period(self, licences: List[PremisesLicence], start: datetime, end: datetime) -> List[PremisesLicence]:
        """Filter licences for a specific time period"""
        filtered = []
        
        for licence in licences:
            # Check if licence falls within the period
            licence_date = licence.granted_date or licence.application_date or licence.scraped_at
            
            if licence_date and start <= licence_date < end:
                filtered.append(licence)
                
        return filtered
    
    def _calculate_summary_stats(self, all_licences: List[PremisesLicence], weekly_licences: List[PremisesLicence]) -> Dict[str, Any]:
        """Calculate summary statistics"""
        stats = {}
        
        # Weekly comparison
        if len(all_licences) > len(weekly_licences):
            previous_week_count = len(all_licences) - len(weekly_licences)  # Simplified
            if previous_week_count > 0:
                stats['weekly_change_percent'] = ((len(weekly_licences) - previous_week_count) / previous_week_count) * 100
            else:
                stats['weekly_change_percent'] = 0
        else:
            stats['weekly_change_percent'] = 0
        
        # Geographic distribution
        stats['unique_postcodes'] = len(set(licence.postcode for licence in weekly_licences if licence.postcode))
        
        # Business insights
        stats['avg_activities_per_licence'] = sum(len(licence.licensable_activities) for licence in weekly_licences) / len(weekly_licences) if weekly_licences else 0
        stats['premises_with_conditions'] = len([l for l in weekly_licences if l.conditions])
        stats['premises_with_hours'] = len([l for l in weekly_licences if l.opening_hours or l.alcohol_hours])
        
        # Risk analysis
        risk_scores = [getattr(licence, 'risk_score', 5) for licence in weekly_licences]
        stats['avg_risk_score'] = sum(risk_scores) / len(risk_scores) if risk_scores else 5
        stats['high_risk_premises'] = len([score for score in risk_scores if score >= 8])
        
        return stats
    
    def _count_by_type(self, licences: List[PremisesLicence]) -> Dict[LicenceType, int]:
        """Count licences by type"""
        counts = {}
        for licence_type in LicenceType:
            counts[licence_type] = len([l for l in licences if l.licence_type == licence_type])
        return counts
    
    def _count_by_status(self, licences: List[PremisesLicence]) -> Dict[LicenceStatus, int]:
        """Count licences by status"""
        counts = {}
        for status in LicenceStatus:
            counts[status] = len([l for l in licences if l.licence_status == status])
        return counts
    
    def _get_top_councils(self, licences: List[PremisesLicence], limit: int = 10) -> List[Dict[str, Any]]:
        """Get top councils by licence count"""
        council_counts = {}
        for licence in licences:
            council_counts[licence.council_name] = council_counts.get(licence.council_name, 0) + 1
        
        sorted_councils = sorted(council_counts.items(), key=lambda x: x[1], reverse=True)
        
        return [
            {'council': council, 'count': count}
            for council, count in sorted_councils[:limit]
        ]
    
    def _create_excel_report(self, report_data: WeeklyReport, all_licences: List[PremisesLicence], weekly_licences: List[PremisesLicence]) -> str:
        """Create Excel report with multiple sheets"""
        filename = f"{self.settings.reports_dir}/weekly_report_{report_data.period_start.strftime('%Y%m%d')}.xlsx"
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Executive Summary Sheet
            self._create_executive_summary_sheet(writer, report_data, weekly_licences)
            
            # New Licences Sheet
            self._create_new_licences_sheet(writer, weekly_licences)
            
            # Analysis Sheets
            self._create_weekly_analysis_sheet(writer, weekly_licences)
            self._create_trends_sheet(writer, all_licences, weekly_licences)
            
        # Apply formatting
        self._format_workbook(filename)
        
        return filename
    
    def _create_executive_summary_sheet(self, writer: pd.ExcelWriter, report_data: WeeklyReport, weekly_licences: List[PremisesLicence]):
        """Create executive summary sheet"""
        summary_data = [
            ['Weekly Report Summary', ''],
            ['Report Period', f"{report_data.period_start.strftime('%Y-%m-%d')} to {report_data.period_end.strftime('%Y-%m-%d')}"],
            ['Generated On', report_data.report_date.strftime('%Y-%m-%d %H:%M')],
            ['', ''],
            ['Key Metrics', ''],
            ['New Licences This Week', report_data.new_licences],
            ['Total Licences in Database', report_data.total_licences],
            ['Councils Monitored', report_data.councils_scraped],
            ['Weekly Change', f"{report_data.summary_stats.get('weekly_change_percent', 0):.1f}%"],
            ['', ''],
            ['Business Intelligence', ''],
            ['Average Activities per Licence', f"{report_data.summary_stats.get('avg_activities_per_licence', 0):.1f}"],
            ['Premises with Operating Conditions', report_data.summary_stats.get('premises_with_conditions', 0)],
            ['High Risk Premises (Score ≥8)', report_data.summary_stats.get('high_risk_premises', 0)],
            ['Average Risk Score', f"{report_data.summary_stats.get('avg_risk_score', 5):.1f}"],
            ['', ''],
            ['Top 5 Councils by New Licences', ''],
        ]
        
        # Add top councils
        for council_data in report_data.top_councils[:5]:
            summary_data.append([council_data['council'], council_data['count']])
            
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    def _create_new_licences_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create new licences data sheet"""
        if not licences:
            pd.DataFrame({'Message': ['No new licences found for this period']}).to_excel(
                writer, sheet_name='New Licences', index=False
            )
            return
            
        # Convert licences to DataFrame
        licence_data = []
        
        for licence in licences:
            licence_data.append({
                'Council': licence.council_name,
                'Premises Name': licence.premises_name,
                'Address': licence.premises_address,
                'Postcode': licence.postcode or '',
                'Business Type': getattr(licence, 'business_type', 'Unknown'),
                'Licence Type': licence.licence_type.value,
                'Status': licence.licence_status.value,
                'Application Date': licence.application_date.strftime('%Y-%m-%d') if licence.application_date else '',
                'Granted Date': licence.granted_date.strftime('%Y-%m-%d') if licence.granted_date else '',
                'Activities': ', '.join(licence.licensable_activities[:3]) + ('...' if len(licence.licensable_activities) > 3 else ''),
                'Risk Score': getattr(licence, 'risk_score', 5),
                'DPS': licence.designated_premises_supervisor or '',
                'Conditions Count': len(licence.conditions),
                'Source URL': str(licence.source_url) if licence.source_url else ''
            })
        
        df = pd.DataFrame(licence_data)
        df.to_excel(writer, sheet_name='New Licences', index=False)
    
    def _create_weekly_analysis_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create weekly analysis with charts data"""
        if not licences:
            return
            
        # Business type analysis
        business_types = {}
        for licence in licences:
            btype = getattr(licence, 'business_type', 'Unknown')
            business_types[btype] = business_types.get(btype, 0) + 1
        
        business_df = pd.DataFrame([
            {'Business Type': btype, 'Count': count, 'Percentage': f"{(count/len(licences)*100):.1f}%"}
            for btype, count in sorted(business_types.items(), key=lambda x: x[1], reverse=True)
        ])
        
        # Activity analysis  
        activity_counts = {}
        for licence in licences:
            for activity in licence.licensable_activities:
                activity_counts[activity] = activity_counts.get(activity, 0) + 1
                
        activity_df = pd.DataFrame([
            {'Activity': activity, 'Count': count, 'Premises': f"{count} premises"}
            for activity, count in sorted(activity_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        ])
        
        # Write to different starting rows
        business_df.to_excel(writer, sheet_name='Weekly Analysis', startrow=0, index=False)
        
        # Add spacing and activity analysis
        start_row = len(business_df) + 3
        pd.DataFrame([[''], ['Top Activities This Week'], ['']]).to_excel(
            writer, sheet_name='Weekly Analysis', startrow=start_row, index=False, header=False
        )
        
        activity_df.to_excel(writer, sheet_name='Weekly Analysis', startrow=start_row + 3, index=False)
    
    def _create_trends_sheet(self, writer: pd.ExcelWriter, all_licences: List[PremisesLicence], weekly_licences: List[PremisesLicence]):
        """Create trends and comparison sheet"""
        # This would contain trend analysis over time
        # For now, create a simple comparison
        
        trends_data = [
            ['Trend Analysis', ''],
            ['', ''],
            ['Metric', 'This Week', 'All Time Average'],
            ['Licences per Week', len(weekly_licences), len(all_licences) // 52 if len(all_licences) > 52 else len(all_licences)],
            ['Average Risk Score', 
             f"{sum(getattr(l, 'risk_score', 5) for l in weekly_licences) / len(weekly_licences):.1f}" if weekly_licences else "0",
             f"{sum(getattr(l, 'risk_score', 5) for l in all_licences) / len(all_licences):.1f}" if all_licences else "0"],
            ['Avg Activities per Licence',
             f"{sum(len(l.licensable_activities) for l in weekly_licences) / len(weekly_licences):.1f}" if weekly_licences else "0",
             f"{sum(len(l.licensable_activities) for l in all_licences) / len(all_licences):.1f}" if all_licences else "0"],
        ]
        
        trends_df = pd.DataFrame(trends_data)
        trends_df.to_excel(writer, sheet_name='Trends', index=False, header=False)
    
    def _create_main_data_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create main data sheet for full dataset report"""
        self._create_new_licences_sheet(writer, licences)  # Same format, different sheet name context
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create summary sheet for full dataset"""
        total_count = len(licences)
        
        # Calculate various metrics
        council_count = len(set(licence.council_name for licence in licences))
        postcode_count = len(set(licence.postcode for licence in licences if licence.postcode))
        
        # Date range
        dates = [licence.granted_date or licence.application_date for licence in licences if licence.granted_date or licence.application_date]
        date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}" if dates else "Unknown"
        
        summary_data = [
            ['Full Dataset Summary', ''],
            ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['', ''],
            ['Overview', ''],
            ['Total Licences', total_count],
            ['Councils Covered', council_count],
            ['Unique Postcodes', postcode_count],
            ['Date Range', date_range],
            ['', ''],
        ]
        
        # Add business type breakdown
        business_types = {}
        for licence in licences:
            btype = getattr(licence, 'business_type', 'Unknown')
            business_types[btype] = business_types.get(btype, 0) + 1
            
        summary_data.append(['Business Type Breakdown', ''])
        for btype, count in sorted(business_types.items(), key=lambda x: x[1], reverse=True):
            summary_data.append([btype, f"{count} ({count/total_count*100:.1f}%)"])
        
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_council_analysis_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create council analysis sheet"""
        council_data = {}
        
        for licence in licences:
            council = licence.council_name
            if council not in council_data:
                council_data[council] = {
                    'Total Licences': 0,
                    'Granted': 0,
                    'Pending': 0,
                    'Business Types': set(),
                    'Avg Risk Score': [],
                    'With Conditions': 0,
                    'Recent Activity': 0
                }
            
            data = council_data[council]
            data['Total Licences'] += 1
            
            if licence.licence_status == LicenceStatus.GRANTED:
                data['Granted'] += 1
            elif licence.licence_status == LicenceStatus.PENDING:
                data['Pending'] += 1
                
            data['Business Types'].add(getattr(licence, 'business_type', 'Unknown'))
            data['Avg Risk Score'].append(getattr(licence, 'risk_score', 5))
            
            if licence.conditions:
                data['With Conditions'] += 1
                
            # Check if recent (last 30 days)
            if licence.granted_date and licence.granted_date > datetime.now() - timedelta(days=30):
                data['Recent Activity'] += 1
        
        # Convert to DataFrame
        council_analysis = []
        for council, data in council_data.items():
            council_analysis.append({
                'Council': council,
                'Total Licences': data['Total Licences'],
                'Granted': data['Granted'],
                'Pending': data['Pending'],
                'Unique Business Types': len(data['Business Types']),
                'Avg Risk Score': f"{sum(data['Avg Risk Score']) / len(data['Avg Risk Score']):.1f}" if data['Avg Risk Score'] else "0",
                'With Conditions': data['With Conditions'],
                'Recent Activity (30d)': data['Recent Activity']
            })
        
        council_df = pd.DataFrame(council_analysis)
        council_df = council_df.sort_values('Total Licences', ascending=False)
        council_df.to_excel(writer, sheet_name='Council Analysis', index=False)
    
    def _create_business_type_analysis_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create business type analysis sheet"""
        business_analysis = {}
        
        for licence in licences:
            btype = getattr(licence, 'business_type', 'Unknown')
            
            if btype not in business_analysis:
                business_analysis[btype] = {
                    'Count': 0,
                    'Avg Risk Score': [],
                    'Common Activities': {},
                    'Avg Conditions': [],
                    'Geographic Spread': set()
                }
            
            data = business_analysis[btype]
            data['Count'] += 1
            data['Avg Risk Score'].append(getattr(licence, 'risk_score', 5))
            data['Avg Conditions'].append(len(licence.conditions))
            
            for activity in licence.licensable_activities:
                data['Common Activities'][activity] = data['Common Activities'].get(activity, 0) + 1
            
            if licence.postcode:
                # Use first part of postcode for geographic spread
                area = licence.postcode.split()[0] if ' ' in licence.postcode else licence.postcode[:3]
                data['Geographic Spread'].add(area)
        
        # Convert to DataFrame
        business_type_analysis = []
        for btype, data in business_analysis.items():
            top_activity = max(data['Common Activities'].items(), key=lambda x: x[1])[0] if data['Common Activities'] else "None"
            
            business_type_analysis.append({
                'Business Type': btype,
                'Count': data['Count'],
                'Percentage': f"{data['Count'] / len(licences) * 100:.1f}%",
                'Avg Risk Score': f"{sum(data['Avg Risk Score']) / len(data['Avg Risk Score']):.1f}" if data['Avg Risk Score'] else "0",
                'Avg Conditions': f"{sum(data['Avg Conditions']) / len(data['Avg Conditions']):.1f}" if data['Avg Conditions'] else "0",
                'Geographic Areas': len(data['Geographic Spread']),
                'Most Common Activity': top_activity
            })
        
        business_df = pd.DataFrame(business_type_analysis)
        business_df = business_df.sort_values('Count', ascending=False)
        business_df.to_excel(writer, sheet_name='Business Type Analysis', index=False)
    
    def _create_activity_analysis_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create licensable activities analysis sheet"""
        activity_analysis = {}
        
        for licence in licences:
            for activity in licence.licensable_activities:
                if activity not in activity_analysis:
                    activity_analysis[activity] = {
                        'Count': 0,
                        'Business Types': {},
                        'Risk Scores': [],
                        'Councils': set()
                    }
                
                data = activity_analysis[activity]
                data['Count'] += 1
                data['Risk Scores'].append(getattr(licence, 'risk_score', 5))
                data['Councils'].add(licence.council_name)
                
                btype = getattr(licence, 'business_type', 'Unknown')
                data['Business Types'][btype] = data['Business Types'].get(btype, 0) + 1
        
        # Convert to DataFrame
        activity_data = []
        for activity, data in activity_analysis.items():
            top_business_type = max(data['Business Types'].items(), key=lambda x: x[1])[0] if data['Business Types'] else "Unknown"
            
            activity_data.append({
                'Activity': activity,
                'Count': data['Count'],
                'Percentage': f"{data['Count'] / len([a for l in licences for a in l.licensable_activities]) * 100:.1f}%",
                'Avg Risk Score': f"{sum(data['Risk Scores']) / len(data['Risk Scores']):.1f}" if data['Risk Scores'] else "0",
                'Councils': len(data['Councils']),
                'Most Common Business Type': top_business_type,
                'Business Types Count': len(data['Business Types'])
            })
        
        activity_df = pd.DataFrame(activity_data)
        activity_df = activity_df.sort_values('Count', ascending=False)
        activity_df.to_excel(writer, sheet_name='Activity Analysis', index=False)
    
    def _create_geographic_analysis_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create geographic analysis sheet by postcode area"""
        geographic_analysis = {}
        
        for licence in licences:
            if not licence.postcode:
                continue
                
            # Extract postcode area (first part)
            area = licence.postcode.split()[0] if ' ' in licence.postcode else licence.postcode[:3]
            
            if area not in geographic_analysis:
                geographic_analysis[area] = {
                    'Count': 0,
                    'Business Types': {},
                    'Activities': set(),
                    'Risk Scores': [],
                    'Councils': set()
                }
            
            data = geographic_analysis[area]
            data['Count'] += 1
            data['Risk Scores'].append(getattr(licence, 'risk_score', 5))
            data['Councils'].add(licence.council_name)
            data['Activities'].update(licence.licensable_activities)
            
            btype = getattr(licence, 'business_type', 'Unknown')
            data['Business Types'][btype] = data['Business Types'].get(btype, 0) + 1
        
        # Convert to DataFrame
        geo_data = []
        for area, data in geographic_analysis.items():
            top_business_type = max(data['Business Types'].items(), key=lambda x: x[1])[0] if data['Business Types'] else "Unknown"
            
            geo_data.append({
                'Postcode Area': area,
                'Licence Count': data['Count'],
                'Avg Risk Score': f"{sum(data['Risk Scores']) / len(data['Risk Scores']):.1f}" if data['Risk Scores'] else "0",
                'Unique Activities': len(data['Activities']),
                'Councils': len(data['Councils']),
                'Most Common Business Type': top_business_type,
                'Business Types Count': len(data['Business Types'])
            })
        
        geo_df = pd.DataFrame(geo_data)
        geo_df = geo_df.sort_values('Licence Count', ascending=False)
        geo_df.to_excel(writer, sheet_name='Geographic Analysis', index=False)
    
    def _create_raw_data_sheet(self, writer: pd.ExcelWriter, licences: List[PremisesLicence]):
        """Create raw data export sheet"""
        raw_data = []
        
        for licence in licences:
            raw_data.append({
                'Licence ID': licence.licence_id,
                'Council': licence.council_name,
                'Council Code': licence.council_code or '',
                'Premises Name': licence.premises_name,
                'Address': licence.premises_address,
                'Postcode': licence.postcode or '',
                'Business Type': getattr(licence, 'business_type', 'Unknown'),
                'Licence Type': licence.licence_type.value,
                'Status': licence.licence_status.value,
                'Application Date': licence.application_date.isoformat() if licence.application_date else '',
                'Granted Date': licence.granted_date.isoformat() if licence.granted_date else '',
                'Effective Date': licence.effective_date.isoformat() if licence.effective_date else '',
                'Activities': '|'.join(licence.licensable_activities),
                'Opening Hours': json.dumps(licence.opening_hours) if licence.opening_hours else '',
                'Alcohol Hours': json.dumps(licence.alcohol_hours) if licence.alcohol_hours else '',
                'DPS': licence.designated_premises_supervisor or '',
                'DPS Licence Number': licence.dps_personal_licence_number or '',
                'Conditions': '|'.join(licence.conditions),
                'Variations': '|'.join(licence.variations),
                'Risk Score': getattr(licence, 'risk_score', 5),
                'Licence Categories': '|'.join(getattr(licence, 'licence_categories', [])),
                'Source URL': str(licence.source_url) if licence.source_url else '',
                'Scraped At': licence.scraped_at.isoformat()
            })
        
        raw_df = pd.DataFrame(raw_data)
        raw_df.to_excel(writer, sheet_name='Raw Data Export', index=False)
    
    def _format_workbook(self, filename: str):
        """Apply formatting to the Excel workbook"""
        try:
            wb = openpyxl.load_workbook(filename)
            
            # Define styles
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            title_font = Font(bold=True, size=14)
            title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # Format each worksheet
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                
                # Format headers (assuming first row contains headers)
                if ws.max_row > 0:
                    for cell in ws[1]:
                        if cell.value:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value:
                            cell.border = thin_border
            
            wb.save(filename)
            logger.info(f"Applied formatting to {filename}")
            
        except Exception as e:
            logger.error(f"Error formatting workbook {filename}: {e}")


# CLI function for standalone usage
def main():
    """Main function for running report generation standalone"""
    import logging
    from .data_processor import DataProcessor
    
    logging.basicConfig(level=logging.INFO)
    
    # Load latest processed data
    data_dir = Path("data/licences")
    if not data_dir.exists():
        print("No processed data found. Run data processing first.")
        return
    
    # Find most recent processed data file
    processed_files = list(data_dir.glob("processed_licences_*.json"))
    if not processed_files:
        print("No processed data files found.")
        return
    
    latest_file = max(processed_files, key=lambda f: f.stat().st_mtime)
    print(f"Loading processed data from: {latest_file}")
    
    # Load processed data
    processor = DataProcessor()
    licences, summary = processor.load_processed_data(str(latest_file))
    
    if not licences:
        print("No licences loaded.")
        return
    
    print(f"Loaded {len(licences)} processed licences")
    
    # Generate reports
    generator = ReportGenerator()
    
    # Generate weekly report
    weekly_report = generator.generate_weekly_report(licences)
    print(f"Weekly report generated: {weekly_report}")
    
    # Generate full dataset report
    full_report = generator.generate_full_dataset_report(licences)
    print(f"Full dataset report generated: {full_report}")


if __name__ == "__main__":
    main()
